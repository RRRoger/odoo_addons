# coding: utf-8

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

try:
    import xlwt
except ImportError:
    xlwt = None

from cStringIO import StringIO
import os


def save_excel(name, excel_data, full_dir, sheet2name=None, sheet2data=None):
    """
        在本地生成excel临时文件, 返回路径待上传
        sheet2name & sheet2data 用来记录报表查询的条件
    """
    wb = openpyxl.Workbook()
    # 激活 worksheet
    ws = wb.active
    ws.title = name
    # 可以附加行，从第一列开始附加
    for r in excel_data:
        ws.append(r)
    ws2 = wb.create_sheet(title=sheet2name)
    for r in sheet2data:
        ws2.append(r)

    # 保存文件
    path = full_dir + '/' + name + '.xlsx'
    wb.save(path)
    return path


def _excel_data_getter_for_openpyxl(name, excel_data, sheet2name=None, sheet2data=None):
    """
        在本地生成excel临时文件, 返回路径待上传
        sheet2name & sheet2data 用来记录报表查询的条件
    """
    wb = openpyxl.Workbook()
    # 激活 worksheet
    ws = wb.active
    ws.title = name
    # 可以附加行，从第一列开始附加
    for r in excel_data:
        ws.append(r)

    if sheet2data:
        ws2 = wb.create_sheet(title=sheet2name)
        for r in sheet2data:
            ws2.append(r)

    fp = StringIO()
    wb.save(fp)
    fp.seek(0)
    data = fp.read()
    fp.close()
    return data


def _excel_data_getter_for_xlsxwriter(name, excel_data, sheet2name=None, sheet2data=None):
    """
        在本地生成excel临时文件, 返回路径待上传
        sheet2name & sheet2data 用来记录报表查询的条件
    """

    title_dict = {'font_name': u'微软雅黑', 'font_size': 12, 'align': 'centre', 'bg_color': "#C8C8C8"}
    row_dict = {'font_name': u'微软雅黑', 'font_size': 10}

    xls = StringIO()
    wb = xlsxwriter.Workbook(xls)
    worksheet = wb.add_worksheet(name)
    title_style = wb.add_format(title_dict)
    row_style = wb.add_format(row_dict)
    _row = 0
    for r in excel_data:
        worksheet.write_row(_row, 0, r, title_style if _row == 0 else row_style)
        _row += 1

    if sheet2data:
        worksheet2 = wb.add_worksheet(sheet2name)
        title_style = wb.add_format(title_dict)
        title_style.set_border(1)
        _row = 0
        for r in sheet2data:
            worksheet2.write_row(_row, 0, r, title_style if _row == 0 else row_style)
            _row += 1

    wb.close()
    xls.seek(0)
    data = xls.getvalue()
    return data


def _excel_data_getter_for_xlwt(name, excel_data, sheet2name=None, sheet2data=None):
    """
        在本地生成excel临时文件, 返回路径待上传
        sheet2name & sheet2data 用来记录报表查询的条件
    """
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet(name)

    # header style
    header_style = xlwt.easyxf('pattern: pattern solid;')

    header_font = xlwt.Font()
    header_font.name = u'微软雅黑'  # 指定“微软雅黑”
    header_font.height = 200
    header_style.font = header_font

    borders = xlwt.Borders()
    borders.bottom = xlwt.Borders.THIN
    header_style.borders = borders

    header_style.pattern.pattern_fore_colour = 0x16

    # normal style
    normal_style = xlwt.XFStyle()

    normal_font = xlwt.Font()
    normal_font.name = u'微软雅黑'  # 指定“微软雅黑”
    normal_font.height = 200

    normal_style.font = normal_font

    # group style
    group_style = xlwt.XFStyle()  # xlwt.easyxf('pattern: pattern solid;')

    group_font = xlwt.Font()
    group_font.name = u'微软雅黑'  # 指定“微软雅黑”
    group_font.height = 200  # 字体大小
    # group_font.colour_index = 0x0C
    # group_font.bold = True
    # group_font.italic = True

    group_style.font = group_font

    # group_style.pattern.pattern_fore_colour = 0x16

    borders = xlwt.Borders()
    borders.bottom = xlwt.Borders.THIN
    group_style.borders = borders

    # 对于group的行，将前面空白的列不应用group样式
    for r in range(len(excel_data)):  # row
        style = normal_style
        flag = 'normal'  # normal, group_no_start, group_start
        if r >= 1:
            flag = 'group_no_start'
        elif r == 0:
            style = header_style

        for c in range(len(excel_data[r])):  # column

            if flag == 'group_no_start':

                if excel_data[r][c] is not None:
                    flag = 'group_start'
                    style = group_style

            if excel_data[r][c] is None:
                worksheet.write(r, c, None, style)
            elif isinstance(excel_data[r][c], (int, long, float, complex)):
                worksheet.write(r, c, excel_data[r][c], style)
            else:
                worksheet.write(r, c, '%s' % excel_data[r][c], style)

    if sheet2name and sheet2data:
        # sheet2 添加, 此处是为了在excel里面添加查询时候的条件
        worksheet2 = workbook.add_sheet(sheet2name)
        for r in range(len(sheet2data)):  # row
            style = normal_style
            flag = 'normal'  # normal, group_no_start, group_start
            if r >= 1:
                flag = 'group_no_start'
            elif r == 0:
                style = header_style

            for c in range(len(sheet2data[r])):  # column

                if flag == 'group_no_start':

                    if sheet2data[r][c] is not None:
                        flag = 'group_start'
                        style = group_style

                if sheet2data[r][c] is None:
                    worksheet2.write(r, c, None, style)
                elif isinstance(sheet2data[r][c], (int, long, float, complex)):
                    worksheet2.write(r, c, sheet2data[r][c], style)
                else:
                    worksheet2.write(r, c, '%s' % sheet2data[r][c], style)

    fp = StringIO()
    workbook.save(fp)
    fp.seek(0)
    data = fp.read()
    fp.close()
    return data


def format_data(headers, data, context=None):
    """
        格式化生成excel所需要的数据格式
        如果有需要求和或者求平均值的 则根据情况处理
    """
    header_keys = []  # excel表头信息
    header_values = []  # 表头对应在数据字典里的key
    sum_index = []  # 需要求总和的index
    avg_index = []  # 需要求平均的index
    _index = 0
    for header in headers:
        header_keys.append(header['name'])
        header_values.append(header['alias'])
        if header['group'] == 'sum':
            sum_index.append(_index)
        elif header['group'] == 'avg':
            avg_index.append(_index)
        _index += 1
    values = []
    for s in data:
        value = []
        for header_value in header_values:
            value.append(s[header_value])
        values.append(value)
    if values:
        if sum_index or avg_index:
            col = zip(*values)
            last_line = ['' for r in header_keys]
            if sum_index:
                for _s in sum_index:
                    last_line[_s] = sum(col[_s])
            if avg_index:
                for _s in sum_index:
                    last_line[_s] = sum(col[_s]) / len(col[_s])
            print last_line
            values.append(last_line)
    return [header_keys] + values


def save_file(full_dir, file_name, base64_data):
    # 保存文件 临时文件
    full_path = os.path.join(full_dir, file_name)
    if not os.path.exists(full_dir):
        os.makedirs(full_dir)
    with open(full_path, 'w') as f:
        f.write(base64_data)
    return full_path


def remove_file(full_path):
    # 删除文件 临时文件
    os.remove(full_path)
    return True


excel_data_getter = _excel_data_getter_for_xlsxwriter


# for lib in LIB_LIST:
#     if eval(lib):
#         excel_data_getter = eval("_excel_data_getter_for_" + lib)
#         break
